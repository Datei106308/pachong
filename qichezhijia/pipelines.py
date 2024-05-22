# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from itemadapter import ItemAdapter
import scrapy
import jsonpath
import json
import openpyxl
from openpyxl.utils import get_column_letter
import os



#使用它来清洗/验证数据、去重和数据保存。
class QichezhijiaPipeline:

    def __init__(self):
        self.f = None

        # 创建空的Workbook对象
        self.wb = openpyxl.Workbook()
        # 创建工作薄
        self.ws = self.wb.active
        # 用append函数往表格添加表头


    def process_item(self, item, spider):
        # 转换后的数据结构列表
        cars_data = []
        cars_data += self.filter_data(item['paramitems'])
        cars_data += self.filter_data(item['configitems'])

        # 遍历数据并写入Excel
        for row in cars_data:
            values = [row['name']] + [row.get(f'car{i + 1}', '') for i in range(len(cars_data[0]) - 1)]
            self.ws.append(values)

            # 自动调整列宽
        for col in self.ws.iter_cols(min_row=1, max_col=self.ws.max_column, max_row=1):
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            if adjusted_width > 25:
                adjusted_width = 25
            self.ws.column_dimensions[get_column_letter(column)].width = adjusted_width

        # 保存Workbook
        self.wb.save("车型数据.xlsx")
        # 如果在Unix或Linux系统上，设置文件权限为775
        if os.name == 'posix':  # 检查是否是Unix或Linux系统
            os.chmod("车型数据.xlsx", 0o775)  # 使用八进制数字设置权限



    #数据处理:根据条件筛选数据
    def filter_data(self,data):
        filter_data = []
        for item in data:
            value_items_values = []
            for value_item in item["valueitems"]:
                if "value" in value_item:
                    value_items_values.append(value_item["value"])
                elif "sublist" in value_item and "subvalue" in value_item["sublist"]:
                    value_items_values.append(value_item["sublist"]["subvalue"])
                    # 如果没有 "value" 也没有 "sublist" 或 "subvalue"，则可以选择忽略或添加默认值

            transformed_item = {
                "name": item["name"],
                "valueitems": value_items_values
            }
            filter_data.append(transformed_item)

        # 构造新的数据结构
        new_data = []
        for item in filter_data:
            new_item = {'name': item['name']}
            for index, value in enumerate(item['valueitems'], start=1):
                new_item[f'car{index}'] = value
            new_data.append(new_item)

        return new_data